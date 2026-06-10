"""
Datadex — Excel Register Parser

Parses Excel register map files (.xlsx) into structured register records.

Supports two formats:
1. IP-XACT hierarchical (e.g. IP-XACT register tables):
   - Metadata block at top: 'name', 'baseAddress', etc. in col B, values in col H
   - Header row: Offset | Bit | Access | INI | ... | Symbol | Description
   - Register group rows: col B = hex offset (e.g. '0004h'), col C = register name
   - Bit-field rows: col B empty, col C = bit range (e.g. '[5:0]')

2. Flat table format:
   - First row with recognisable column headers (name, address, bit_field, etc.)
   - One register per row
"""

import os
import re
from dataclasses import dataclass
from typing import List, Optional

import openpyxl


@dataclass
class Register:
    """A single hardware register entry."""
    name: str = ""
    address: str = ""
    bit_field: str = ""
    access: str = ""
    reset_value: str = ""
    description: str = ""
    source: str = ""


# Column name aliases for flat-table auto-detection
_COLUMN_PATTERNS = {
    "name": ["register", "register name", "reg_name", "regname", "name", "register_name", "signal", "peripheral"],
    "address": ["address", "addr", "offset", "base", "address offset", "reg_addr", "register address"],
    "bit_field": ["bit", "bits", "bit field", "bit_field", "field", "bit range", "bit-range"],
    "access": ["access", "type", "access type", "rw", "direction", "permission", "access_type", "r/w", "read/write"],
    "reset_value": ["reset", "reset value", "default", "reset_value", "power-on", "por value", "init value"],
    "description": ["description", "desc", "function", "description/function", "detail", "comment", "remarks", "note"],
}

_HEX_OFFSET_RE = re.compile(r"^[0-9A-Fa-f]+[hH]?$")
_BIT_RANGE_RE = re.compile(r"^\[[\d:]+\]$")


def _normalize(s: str) -> str:
    return s.strip().lower().replace(" ", "_").replace("/", "_").replace("-", "_").replace(".", "_")


def _cell(row, idx) -> str:
    """Safely get a cell value as a stripped string."""
    if idx < len(row) and row[idx] is not None:
        return str(row[idx]).strip()
    return ""


def _is_ipxact_format(rows) -> bool:
    """Return True if the sheet uses the IP-XACT hierarchical layout."""
    for row in rows[:30]:
        if row and len(row) > 1 and row[1] is not None and str(row[1]).strip() == "baseAddress":
            return True
    return False


def _parse_ipxact_sheet(rows, sheet_name: str, source_name: str) -> List[Register]:
    """Parse one sheet in the IP-XACT hierarchical register format."""
    registers: List[Register] = []

    # --- extract peripheral metadata (name + base address) ---
    base_address = 0
    for row in rows[:30]:
        key = _cell(row, 1)
        val = _cell(row, 7)
        if key == "baseAddress" and val:
            try:
                base_address = int(val, 16)
            except (ValueError, TypeError):
                pass

    # --- find data header row (col B == 'Offset', col C == 'Bit') ---
    header_row_idx = None
    for i, row in enumerate(rows):
        if _cell(row, 1).lower() == "offset" and _cell(row, 2).lower() == "bit":
            header_row_idx = i
            break

    if header_row_idx is None:
        return registers

    # --- parse register entries ---
    current_name = ""
    current_abs_addr = base_address

    for row in rows[header_row_idx + 1:]:
        col_b = _cell(row, 1)   # Offset (hex) or empty
        col_c = _cell(row, 2)   # Register name or bit range
        col_d = _cell(row, 3)   # Access
        col_e = _cell(row, 4)   # INI / reset value
        col_h = _cell(row, 7)   # Symbol
        col_i = _cell(row, 8)   # Description

        if col_b:
            # New register group row — col_b is the offset, col_c is the register name
            offset_str = col_b.rstrip("hH")
            try:
                offset = int(offset_str, 16)
            except ValueError:
                continue
            current_name = col_c
            current_abs_addr = base_address + offset

        elif col_c and _BIT_RANGE_RE.match(col_c):
            # Bit-field row — skip reserved fields
            if col_h in ("--", "") or col_h.upper() in ("RSVD", "RESERVED"):
                continue
            if not current_name:
                continue

            desc = f"{col_h}: {col_i}" if col_h and col_i else (col_h or col_i)

            registers.append(Register(
                name=current_name,
                address=hex(current_abs_addr),
                bit_field=col_c,
                access=col_d,
                reset_value=col_e,
                description=desc,
                source=source_name,
            ))

    return registers


# ── flat-table helpers (legacy format) ───────────────────────────────────────

def _detect_column_mapping(headers: List[str]) -> dict:
    mapping = {}
    for col_idx, header in enumerate(headers):
        norm = _normalize(header)
        for field_name, patterns in _COLUMN_PATTERNS.items():
            if any(_normalize(p) == norm or norm.startswith(_normalize(p)) for p in patterns):
                mapping[field_name] = col_idx
                break
    return mapping


def _parse_flat_sheet(rows, source_name: str) -> List[Register]:
    """Parse one sheet in the flat-table register format."""
    registers: List[Register] = []

    # Find header row — must have 'name' or 'address' columns
    header_row_idx = None
    for i, row in enumerate(rows[:30]):
        row_text = [str(c).strip().lower() if c else "" for c in row]
        if any(_normalize(c) in _COLUMN_PATTERNS["name"] or
               _normalize(c) in _COLUMN_PATTERNS["address"] for c in row_text if c):
            # Make sure this is not a key-value metadata cell (IP-XACT style)
            non_empty = [c for c in row_text if c]
            if len(non_empty) >= 3:  # real header rows have several non-empty cells
                header_row_idx = i
                break

    if header_row_idx is None:
        return registers

    headers = [str(c).strip() if c else "" for c in rows[header_row_idx]]
    mapping = _detect_column_mapping(headers)

    if "name" not in mapping and "address" not in mapping:
        return registers

    for row in rows[header_row_idx + 1:]:
        row_vals = [str(c).strip() if c else "" for c in row]
        if not any(row_vals):
            continue

        reg = Register(source=source_name)
        if "name" in mapping:
            reg.name = row_vals[mapping["name"]] if mapping["name"] < len(row_vals) else ""
        if "address" in mapping:
            reg.address = row_vals[mapping["address"]] if mapping["address"] < len(row_vals) else ""
        if "bit_field" in mapping:
            reg.bit_field = row_vals[mapping["bit_field"]] if mapping["bit_field"] < len(row_vals) else ""
        if "access" in mapping:
            reg.access = row_vals[mapping["access"]] if mapping["access"] < len(row_vals) else ""
        if "reset_value" in mapping:
            reg.reset_value = row_vals[mapping["reset_value"]] if mapping["reset_value"] < len(row_vals) else ""
        if "description" in mapping:
            reg.description = row_vals[mapping["description"]] if mapping["description"] < len(row_vals) else ""

        if not reg.name and not reg.address:
            continue

        registers.append(reg)

    return registers


# ── public API ────────────────────────────────────────────────────────────────

class ExcelRegisterParser:
    """Parse .xlsx register map files into structured Register records."""

    def __init__(self, sheet_name: Optional[str] = None):
        self.sheet_name = sheet_name

    def parse_file(self, filepath: str, workspace: str = "") -> List[Register]:
        """Parse an Excel register file into structured records."""
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File not found: {filepath}")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        source_name = os.path.basename(filepath)

        if self.sheet_name:
            if self.sheet_name not in wb.sheetnames:
                wb.close()
                raise ValueError(f"Sheet '{self.sheet_name}' not found in {source_name}")
            sheets = [(self.sheet_name, wb[self.sheet_name])]
        else:
            sheets = [(ws.title, ws) for ws in wb.worksheets]

        all_registers: List[Register] = []
        for sheet_name, ws in sheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            if _is_ipxact_format(rows):
                regs = _parse_ipxact_sheet(rows, sheet_name, source_name)
            else:
                regs = _parse_flat_sheet(rows, source_name)

            all_registers.extend(regs)

        wb.close()
        return all_registers
